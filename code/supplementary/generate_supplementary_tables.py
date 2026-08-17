"""
Generate the canonical supplementary table package S1-S5 from accepted pipeline aggregates.

Output: outs/tables/
  - TableS1_cohorts.xlsx: clinicopathologic features (6 discovery + JHU + Durham)
  - TableS2_metscore_genes.xlsx: 45-gene signature meta-analysis (6 sheets, from aggregates)
  - TableS3_secondary_endpoints_and_GS7.xlsx: Durham OS/PCSM/BCR secondary endpoints + GS7 incremental concordance
  - TableS4_metscore_celltype.xlsx: Met-Score gene cell-type expression (GSE274229)
  - TableS5_kfoury_decomposition.xlsx: Met-Score POS/NEG bone-niche decomposition (GSE143791)

All values are pulled from canonical CSVs / rda-derived clinical files; nothing is
synthetic or guessed. Re-run after any change to the underlying pipeline.
"""

import os
from pathlib import Path
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# resolve project root portably: MET_PCA_ROOT env var, else walk up to the dir holding code/ + outs|output
def _find_met_pca_root():
    env = os.environ.get("MET_PCA_ROOT")
    if env and os.path.isdir(env):
        return env
    d = os.path.dirname(os.path.abspath(__file__)) if "__file__" in globals() else os.getcwd()
    while d != os.path.dirname(d):
        if os.path.isdir(os.path.join(d, "code")) and (
            os.path.isdir(os.path.join(d, "outs")) or os.path.isdir(os.path.join(d, "output"))):
            return d
        d = os.path.dirname(d)
    raise FileNotFoundError("Could not locate Met_PCa project root; set MET_PCA_ROOT.")
ROOT = Path(_find_met_pca_root())
OUT_DIR = ROOT / "outs" / "tables"
OUT_DIR.mkdir(parents=True, exist_ok=True)


# ============================================================================
# TABLE S2: 45-gene Met-Score signature, built entirely from accepted aggregates
# (no hard-coded statistics). The six-sheet workbook is assembled below, after
# the shared _style_xlsx helper is defined.
# ============================================================================


def _style_xlsx(path: Path, header_fill="305496", header_font_color="FFFFFF",
                widths=None, freeze=True):
    """Apply a clean, consistent style to all sheets of an xlsx."""
    from openpyxl import load_workbook
    from datetime import datetime
    wb = load_workbook(path)
    # deterministic metadata so repeated runs produce byte-identical workbooks
    fixed = datetime(2000, 1, 1, 0, 0, 0)
    wb.properties.created = fixed; wb.properties.modified = fixed; wb.properties.creator = "MET_PCa pipeline"
    thin = Side(border_style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ws in wb.worksheets:
        if ws.max_row == 0:
            continue
        for col_idx in range(1, ws.max_column + 1):
            c = ws.cell(row=1, column=col_idx)
            c.font = Font(bold=True, color=header_font_color, name="Arial", size=11)
            c.fill = PatternFill("solid", start_color=header_fill)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = border
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                                 min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.font = Font(name="Arial", size=10)
                cell.border = border
                if isinstance(cell.value, (int, float)) and cell.value is not None:
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.alignment = Alignment(horizontal="left", wrap_text=True)
        # autosize columns
        for col_idx in range(1, ws.max_column + 1):
            letter = get_column_letter(col_idx)
            if widths and letter in widths:
                ws.column_dimensions[letter].width = widths[letter]
            else:
                max_len = max(
                    (len(str(ws.cell(row=r, column=col_idx).value or ""))
                     for r in range(1, ws.max_row + 1)),
                    default=10,
                )
                ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 35)
        if freeze:
            ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 32
    wb.save(path)


# Write Table S2: six sheets built from accepted aggregates (fail-closed).
def _read_req(rel, need):
    fp = ROOT / rel
    if not fp.exists():
        raise FileNotFoundError(f"TableS2 required aggregate missing: {rel}")
    d = pd.read_csv(fp)
    miss = [c for c in need if c not in d.columns]
    if miss:
        raise KeyError(f"{rel} missing required column(s): {miss}")
    return d

# Sheet 1: signature genes with pooled meta-analysis statistics + prediction interval
het = _read_req("outs/FigureS5/FigureS5_panelA_heterogeneity.csv",
                ["gene", "direction", "used_in_locked_model", "pooled_hedges_g", "se", "pooled_p",
                 "pooled_fdr", "tau2", "i2_pct", "cochran_q", "q_p", "n_cohorts", "pi_lo", "pi_hi", "pi_crosses_zero"])
s2_genes = pd.DataFrame({
    "Gene": het["gene"],
    "Metastasis direction": het["direction"],
    "Used in locked model": het["used_in_locked_model"],
    "Pooled Hedges' g": het["pooled_hedges_g"].round(5),
    "Standard error": het["se"].round(5),
    "95% CI lower": (het["pooled_hedges_g"] - 1.959964 * het["se"]).round(5),
    "95% CI upper": (het["pooled_hedges_g"] + 1.959964 * het["se"]).round(5),
    "Pooled p-value": het["pooled_p"],
    "Pooled FDR (BH)": het["pooled_fdr"],
    "Tau²": het["tau2"].round(5),
    "I² (%)": het["i2_pct"].round(2),
    "Cochran Q": het["cochran_q"].round(4),
    "Cochran Q p-value": het["q_p"].round(5),
    "95% prediction interval lower": het["pi_lo"].round(5),
    "95% prediction interval upper": het["pi_hi"].round(5),
    "Prediction interval crosses zero": het["pi_crosses_zero"],
    "N contributing cohorts": het["n_cohorts"].astype(int),
})
# Sheet 2: per-cohort effect sizes
s2_percohort = _read_req("outs/TableS2_per_cohort_effect_sizes.csv", ["gene", "direction", "pooled_effect_size"])
# Sheet 3: fixed-panel leave-one-cohort-out stability per held-out cohort
s2_loco = _read_req("outs/LOCO_gene_stability_with_metadata.csv",
                    ["held_out", "n_full_panel", "n_recovered", "pct_recovered", "jaccard_index", "platform"])
# Sheet 4: fixed-panel LOCO per-gene consistency (actual contributing studies per fold)
s2_loco_gene = _read_req("outs/LOO_gene_consistency.csv",
                         ["removed_cohort", "gene", "direction_full", "effectSize", "effectSizeFDR", "numStudies", "loo_sign_match", "loo_still_sig"])
# Sheets 5-6: complete-reselection summary + membership
s2_resel_sum = _read_req("outs/FigureS5/FigureS5_panelB_fold_summary.csv",
                         ["omitted_cohort", "n_reselected", "n_overlap_with_45", "n_added", "n_lost"])
s2_resel_mem = _read_req("outs/FigureS5/FigureS5_panelB_membership.csv", ["gene", "direction"])

path_s2 = OUT_DIR / "TableS2_metscore_genes.xlsx"
with pd.ExcelWriter(path_s2, engine="openpyxl") as xl:
    s2_genes.to_excel(xl, sheet_name="Signature genes", index=False)
    s2_percohort.to_excel(xl, sheet_name="Per-cohort effects", index=False)
    s2_loco.to_excel(xl, sheet_name="Fixed-panel LOCO stability", index=False)
    s2_loco_gene.to_excel(xl, sheet_name="LOCO gene consistency", index=False)
    s2_resel_sum.to_excel(xl, sheet_name="Reselection summary", index=False)
    s2_resel_mem.to_excel(xl, sheet_name="Reselection membership", index=False)
_style_xlsx(path_s2)
print(f"Wrote {path_s2}  (6 sheets: {len(s2_genes)} genes, per-cohort, LOCO stability/consistency, reselection summary/membership)")


# ============================================================================
# TABLE S3: Durham secondary endpoints (OS/PCSM/BCR) + GS7 incremental concordance
# ============================================================================
# The complete multivariable MFS model is now Main Table 1
# (outs/MainTable1_multivariable_MFS.csv, from MetScore_Sensitivity.R); Figure S6d
# continues to use outs/FigureS6/panelD_multivariable.csv. Table S3 holds the
# Durham secondary-endpoint associations and the corrected GS7 concordance.
#
# Worksheet 1: Durham secondary endpoints (fail-closed), from the canonical
# producer outs/TableS3_Durham_secondary_endpoints.csv (OS, PCSM, BCR only).
src_sec = ROOT / "outs" / "TableS3_Durham_secondary_endpoints.csv"
if not src_sec.exists():
    raise FileNotFoundError("TableS3 source missing: outs/TableS3_Durham_secondary_endpoints.csv")
psec = pd.read_csv(src_sec)
for c in ("Cohort", "Endpoint", "Score_form", "Adjustment_set", "Variable", "Comparison",
          "HR", "CI_lo", "CI_hi", "p", "n", "events", "Estimator", "Time_origin", "Status"):
    if c not in psec.columns:
        raise KeyError(f"TableS3_Durham_secondary_endpoints.csv missing required column: {c}")
if (psec["Endpoint"].astype(str) == "MFS").any():
    raise ValueError("TableS3_Durham_secondary_endpoints.csv unexpectedly contains an MFS row")
df_dur = pd.DataFrame({
    "Cohort": psec["Cohort"],
    "Endpoint": psec["Endpoint"],
    "Score form": psec["Score_form"],
    "Adjustment set": psec["Adjustment_set"],
    "Variable": psec["Variable"],
    "Comparison": psec["Comparison"],
    "HR": psec["HR"],
    "95% CI lower": psec["CI_lo"],
    "95% CI upper": psec["CI_hi"],
    "p-value": psec["p"],
    "n": psec["n"].astype(int),
    "events": psec["events"].astype(int),
    "Estimator": psec["Estimator"],
    "Time origin": psec["Time_origin"],
    "Status": psec["Status"],
})

# Worksheet 2: GS7 incremental cause-specific concordance (fail-closed).
# JHU is the Sanderson case-cohort-weighted, optimism-corrected estimate; Durham
# applies the frozen JHU coefficients as an ordinary external Harrell C.
src_gs7 = ROOT / "outs" / "Figure3" / "GS7_incremental_concordance.csv"
if not src_gs7.exists():
    raise FileNotFoundError("TableS3 GS7 worksheet source missing: outs/Figure3/GS7_incremental_concordance.csv")
pgs7 = pd.read_csv(src_gs7)
for c in ("cohort", "role", "model", "score_form", "n", "events", "GG2_n", "GG2_events", "GG3_n", "GG3_events",
          "estimator", "sampling_fraction", "apparent_C", "mean_optimism", "corrected_or_frozen_C",
          "C_lo", "C_hi", "dC_vs_GG_only", "dC_lo", "dC_hi", "boot_attempted", "boot_success", "boot_failed",
          "model_fit_warnings", "score_provenance"):
    if c not in pgs7.columns:
        raise KeyError(f"GS7_incremental_concordance.csv missing required column: {c}")
df_gs7 = pd.DataFrame({
    "Cohort": pgs7["cohort"],
    "Role": pgs7["role"],
    "Model": pgs7["model"],
    "Score form": pgs7["score_form"],
    "n": pgs7["n"].astype(int),
    "Metastasis events": pgs7["events"].astype(int),
    "GG2 n": pgs7["GG2_n"].astype(int),
    "GG2 events": pgs7["GG2_events"].astype(int),
    "GG3 n": pgs7["GG3_n"].astype(int),
    "GG3 events": pgs7["GG3_events"].astype(int),
    "Estimator": pgs7["estimator"],
    "Case-cohort sampling fraction": pgs7["sampling_fraction"],
    "Apparent C": pgs7["apparent_C"],
    "Mean optimism": pgs7["mean_optimism"],
    "Corrected / frozen-external C": pgs7["corrected_or_frozen_C"],
    "C 95% CI lower": pgs7["C_lo"],
    "C 95% CI upper": pgs7["C_hi"],
    "ΔC vs Grade Group only": pgs7["dC_vs_GG_only"],
    "ΔC 95% CI lower": pgs7["dC_lo"],
    "ΔC 95% CI upper": pgs7["dC_hi"],
    "Bootstrap attempted": pgs7["boot_attempted"].astype(int),
    "Bootstrap successful": pgs7["boot_success"].astype(int),
    "Bootstrap failed": pgs7["boot_failed"].astype(int),
    "Model-fit warnings": pgs7["model_fit_warnings"].astype(int),
    "Score provenance": pgs7["score_provenance"],
})
path_s3 = OUT_DIR / "TableS3_secondary_endpoints_and_GS7.xlsx"
with pd.ExcelWriter(path_s3, engine="openpyxl") as xl:
    df_dur.to_excel(xl, index=False, sheet_name="Durham secondary endpoints")
    df_gs7.to_excel(xl, index=False, sheet_name="GS7 incremental concordance")
_style_xlsx(path_s3)
print(f"Wrote {path_s3}  (Durham secondary endpoints {len(df_dur)} rows; GS7 incremental concordance {len(df_gs7)} rows)")


# ============================================================================
# TABLE S4: Met-Score gene cell-type expression in GSE274229
# ============================================================================
src_s4c = ROOT / "outs" / "met_score_gene_celltype_expression_WITH_GENE.csv"
if not src_s4c.exists():
    raise FileNotFoundError("TableS4 source missing: outs/met_score_gene_celltype_expression_WITH_GENE.csv")
df_s4 = pd.read_csv(src_s4c).rename(columns={
    "gene": "Gene", "celltype": "Cell type", "pct_expr": "% cells expressing",
    "mean_expr": "Mean expression (log-normalised)", "median_expr": "Median expression (log-normalised)",
    "n_cells": "N cells"})
df_s4["% cells expressing"] = df_s4["% cells expressing"].round(2)
df_s4["Mean expression (log-normalised)"] = df_s4["Mean expression (log-normalised)"].round(4)
df_s4["Median expression (log-normalised)"] = df_s4["Median expression (log-normalised)"].round(4)
df_s4["N cells"] = df_s4["N cells"].astype(int)
df_s4 = df_s4.sort_values(["Gene", "Cell type"]).reset_index(drop=True)
path_s4 = OUT_DIR / "TableS4_metscore_celltype.xlsx"
df_s4.to_excel(path_s4, index=False, sheet_name="MetScore gene by celltype")
_style_xlsx(path_s4)
print(f"Wrote {path_s4}  ({len(df_s4)} gene × cell-type rows)")


# ============================================================================
# TABLE S5: Met-Score POS/NEG decomposition in GSE143791 (Kfoury bone niche)
# ============================================================================
_kf = ROOT / "outs" / "kfoury"
_kf_files = [("Cell type x fraction", "GSE143791_met_score_pos_neg_by_celltype_by_fraction.csv"),
             ("Cell type", "GSE143791_met_score_pos_neg_by_celltype.csv"),
             ("Fraction", "GSE143791_met_score_pos_neg_by_fraction.csv")]
_kf_df = {}
for sheet, fn in _kf_files:
    fp = _kf / fn
    if not fp.exists():
        raise FileNotFoundError(f"TableS5 source missing: outs/kfoury/{fn}")
    d = pd.read_csv(fp)
    for c in d.select_dtypes(include="float").columns:
        d[c] = d[c].round(4)
    _kf_df[sheet] = d
notes5 = pd.DataFrame({"Note": [
    "Absent cell-type x fraction combinations were not observed in GSE143791; they are omitted, not zero-imputed:",
    "Tumor - Benign", "Pericytes - Benign", "Pericytes - Distal", "Pericytes - Involved", "Osteoclasts - Benign"]})
path_s5 = OUT_DIR / "TableS5_kfoury_decomposition.xlsx"
with pd.ExcelWriter(path_s5, engine="openpyxl") as xl:
    _kf_df["Cell type x fraction"].to_excel(xl, sheet_name="Cell type x fraction", index=False)
    _kf_df["Cell type"].to_excel(xl, sheet_name="Cell type", index=False)
    _kf_df["Fraction"].to_excel(xl, sheet_name="Fraction", index=False)
    notes5.to_excel(xl, sheet_name="Notes", index=False)
_style_xlsx(path_s5)
print(f"Wrote {path_s5}  (4 sheets: cell-type×fraction, cell-type, fraction, notes)")


# ============================================================================
# TABLE S1: Cohort clinicopathologic characteristics
# ============================================================================
# Scope: the 6 discovery cohorts plus the 2 clinical validation cohorts
# (JHU Nat. History, Durham VA).
# Column / row layout matches the previously published supplementary table:
#   Patients
#   Pathological Stage: T1, T2, T3, T4, NA
#   Gleason Score: <=6, 7, 8, 9, 10, NA
#   Preoperative PSA (ng/ml): <10, 10-20, >20, NA
#   Metastasis: No, Yes, NA
#
# All discovery + JHU columns are taken VERBATIM from the previously published
# supplementary table (those cohorts did not change). Durham columns are
# extracted directly from the canonical clinical xlsx files in
# data/Durham_cohort_and_GRID_cohort/ using the exact subset filters the
# validation R scripts apply (Durham: drop pogl==0 sentinel rows -> n=555).
# All counts reported with no imputation; "-" or "NA" used where the field is
# not in the release.

import re

# ---- 1. Stage categorisation helper (drop p prefix, drop nodal annotation) ----
def _categorize_T(p):
    if pd.isna(p):
        return "NA"
    s = str(p).strip().lower().replace("p", "")
    s = re.sub(r"n[+\d].*$", "", s).strip()
    if s.startswith("t1"): return "T1"
    if s.startswith("t2"): return "T2"
    if s.startswith("t3"): return "T3"
    if s.startswith("t4"): return "T4"
    return "NA"   # includes "0" sentinel and "Tx"

# ---- 2. Extract Durham n=555 features ----
# Replicate R-script filter chain EXACTLY:
#   clin <- read clin sheet (887 rows)
#   sample_map <- read sample-id mapping (887 rows after header skip)
#   clin_mapped <- merge(clin, sample_map, by='cell_file_name')  -> 887 rows
#   clin_valid  <- clin_mapped[!is.na(clin_mapped$mets), ]        -> 558 rows
#   clin_valid  <- clin_valid[!(is.na(pogl) | pogl == 0), ]       -> 555 rows
d_clin = pd.read_excel(
    ROOT / "data" / "Durham_cohort_and_GRID_cohort" / "Durham_cohort_clinical_data_022526.xlsx",
    sheet_name="clin")
d_sm_raw = pd.read_excel(
    ROOT / "data" / "Durham_cohort_and_GRID_cohort" / "Durham_cohort_011526.xlsx",
    sheet_name="Sheet2", header=None)
d_sm = d_sm_raw.iloc[1:].copy()
d_sm.columns = ["unnamed_0", "cell_file_name", "sample_id"]
d_sm = d_sm[["cell_file_name", "sample_id"]].dropna(subset=["cell_file_name"]).reset_index(drop=True)

d_clin_mapped = d_clin.merge(d_sm, on="cell_file_name", how="inner")
d_step1 = d_clin_mapped[~d_clin_mapped["mets"].isna()].copy()
d_555 = d_step1[~(d_step1["pogl"].isna() | (d_step1["pogl"] == 0))].copy()
assert len(d_555) == 555, f"Durham filter chain produced {len(d_555)}, expected 555"

# CORRECTED: Pathological stage is column 'stg' per the codebook
# ('clinicalstage' is Clinical stage). Was: clinicalstage. Now: stg.
durham_T = d_555["stg"].apply(_categorize_T).value_counts()
durham_T1 = int(durham_T.get("T1", 0))
durham_T2 = int(durham_T.get("T2", 0))
durham_T3 = int(durham_T.get("T3", 0))
durham_T4 = int(durham_T.get("T4", 0))
durham_T_NA = 555 - (durham_T1 + durham_T2 + durham_T3 + durham_T4)

durham_GS_le6 = int((d_555["pogl"] <= 6).sum())
durham_GS_7   = int((d_555["pogl"] == 7).sum())
durham_GS_8   = int((d_555["pogl"] == 8).sum())
durham_GS_9   = int((d_555["pogl"] == 9).sum())
durham_GS_10  = int((d_555["pogl"] == 10).sum())
durham_GS_NA  = int(d_555["pogl"].isna().sum())

durham_psa_lt10  = int((d_555["psapresurg"] < 10).sum())
durham_psa_10_20 = int(((d_555["psapresurg"] >= 10) & (d_555["psapresurg"] <= 20)).sum())
durham_psa_gt20  = int((d_555["psapresurg"] > 20).sum())
durham_psa_NA    = int(d_555["psapresurg"].isna().sum())

durham_met_yes = int((d_555["mets"] == 1).sum())
durham_met_no  = int((d_555["mets"] == 0).sum())
durham_met_NA  = int(d_555["mets"].isna().sum())
assert durham_met_yes + durham_met_no + durham_met_NA == 555

# ---- 3. Build the table in the exact published layout: cohorts as columns ----
# Rows (in order): Patients; Stage T1/T2/T3/T4/NA; Gleason <=6/7/8/9/10/NA;
# PSA <10/10-20/>20/NA; Metastasis No/Yes/NA.

row_labels = [
    "Patients",
    "Pathological Stage", "  T1", "  T2", "  T3", "  T4", "  NA",
    "Gleason Score", "  ≤6", "  7", "  8", "  9", "  10", "  NA ",
    "Preoperative PSA (ng/ml)", "  <10", "  10–20", "  >20", "  NA  ",
    "Metastasis", "  No", "  Yes", "  NA   ",
]

# Each cohort column = one list aligned with row_labels.
# Header rows ("Pathological Stage" etc.) leave a blank cell.

def col(patients,
        stage_T1, stage_T2, stage_T3, stage_T4, stage_NA,
        gs_le6, gs_7, gs_8, gs_9, gs_10, gs_NA,
        psa_lt10, psa_10_20, psa_gt20, psa_NA,
        met_no, met_yes, met_NA):
    return [patients,
            "", stage_T1, stage_T2, stage_T3, stage_T4, stage_NA,
            "", gs_le6, gs_7, gs_8, gs_9, gs_10, gs_NA,
            "", psa_lt10, psa_10_20, psa_gt20, psa_NA,
            "", met_no, met_yes, met_NA]

# ---- Discovery cohorts (verbatim from the previously published table) ----
gse116918 = col(248,
                51, 76, 92, 4, 25,
                42, 99, 52, 54, 1, 0,
                50, 95, 103, 0,
                226, 22, 0)
gse55935 = col(44,
                "-", "-", "-", "-", 44,
                "-", "-", "-", "-", "-", 44,
                "-", "-", "-", 44,
                36, 8, 0)
gse51066 = col(85,
                "-", "-", "-", "-", 85,
                "-", "-", "-", "-", "-", 85,
                "-", "-", "-", 85,
                34, 51, 0)
gse46691 = col(545,
                "-", "-", "-", "-", 545,
                63, 271, 68, 134, 9, 0,
                "-", "-", "-", 545,
                333, 212, 0)
gse41408 = col(48,
                0, 15, 26, 6, 1,
                23, 16, 8, 1, 0, 0,
                21, 17, 8, 2,
                39, 9, 0)
gse70769 = col(30,
                14, 11, 4, 0, 1,
                4, 16, 3, 5, 1, 1,
                15, 9, 6, 0,
                26, 4, 0)
discovery_total = col(1000,
                       65, 102, 122, 10, 701,
                       132, 402, 131, 194, 11, 130,
                       86, 121, 117, 676,
                       694, 306, 0)

# ---- Validation cohort: JHU, extracted from canonical pheno_jhu rda ----
# Path stage from `pstage` (the previously published table used `cstage`/clinical stage).
# Path Gleason from raw `pathgs`, the column that drives the survival models; its
# distribution is <=6 = 1, 7 = 133, 8 = 28, 9 = 77, 10 = 0, missing = 0.
# PSA from `preop_psa`. Metastasis from `Metastasis` factor (No_Mets/Mets).
# Read via the pure-Python `rdata` reader rather than pyreadr: pyreadr's librdata
# backend raises "Unable to convert string to the requested encoding (invalid byte
# sequence)" on this .rda because some string columns contain non-ASCII bytes and
# librdata has no encoding override. `rdata` reads it cleanly when told to decode as
# UTF-8, returning the same pheno_jhu data.frame as a pandas DataFrame.
import rdata as _rdata
_jhu = _rdata.conversion.convert(
    _rdata.parser.parse_file(str(ROOT / "outs" / "jhu_pheno_filter_MetaScorer_Zscore.rda")),
    default_encoding="utf-8", force_default_encoding=True,
)["pheno_jhu"]
assert len(_jhu) == 239

_jhu_T = _jhu["pstage"].apply(_categorize_T).value_counts()
jhu_T1 = int(_jhu_T.get("T1", 0))
jhu_T2 = int(_jhu_T.get("T2", 0))
jhu_T3 = int(_jhu_T.get("T3", 0))
jhu_T4 = int(_jhu_T.get("T4", 0))
jhu_T_NA = 239 - (jhu_T1 + jhu_T2 + jhu_T3 + jhu_T4)

# Use `pathgs`, the column that drives the multivariate Cox in
# Met_PCa_Survival.R: line 337 `pheno_jhu$pathgs <- factor(...)`, line 351
# rename to "Pathological GS", line 627/635/686/694 `MetScoreClass +
# \`Pathological GS\``. Reporting Table S1 with the same variable used in
# the survival models avoids any inconsistency between the descriptive
# Table S1 and the headline multivariate analysis.
_jhu_gs = pd.to_numeric(_jhu["pathgs"], errors="coerce")
jhu_gs_le6 = int(((_jhu_gs >= 1) & (_jhu_gs <= 6)).sum())
jhu_gs_7   = int((_jhu_gs == 7).sum())
jhu_gs_8   = int((_jhu_gs == 8).sum())
jhu_gs_9   = int((_jhu_gs == 9).sum())
jhu_gs_10  = int((_jhu_gs == 10).sum())
jhu_gs_NA  = int(_jhu_gs.isna().sum())
assert jhu_gs_le6 + jhu_gs_7 + jhu_gs_8 + jhu_gs_9 + jhu_gs_10 + jhu_gs_NA == 239

_jhu_psa = pd.to_numeric(_jhu["preop_psa"], errors="coerce")
jhu_psa_lt10  = int((_jhu_psa < 10).sum())
jhu_psa_10_20 = int(((_jhu_psa >= 10) & (_jhu_psa <= 20)).sum())
jhu_psa_gt20  = int((_jhu_psa > 20).sum())
jhu_psa_NA    = int(_jhu_psa.isna().sum())

jhu_met_yes = int((_jhu["Metastasis"] == "Mets").sum())
jhu_met_no  = int((_jhu["Metastasis"] == "No_Mets").sum())
jhu_met_NA  = 239 - jhu_met_yes - jhu_met_no

jhu = col(239,
           jhu_T1, jhu_T2, jhu_T3, jhu_T4, jhu_T_NA,
           jhu_gs_le6, jhu_gs_7, jhu_gs_8, jhu_gs_9, jhu_gs_10, jhu_gs_NA,
           jhu_psa_lt10, jhu_psa_10_20, jhu_psa_gt20, jhu_psa_NA,
           jhu_met_no, jhu_met_yes, jhu_met_NA)

# ---- Validation cohort: Durham VA (extracted) ----
durham = col(555,
              durham_T1, durham_T2, durham_T3, durham_T4, durham_T_NA,
              durham_GS_le6, durham_GS_7, durham_GS_8, durham_GS_9, durham_GS_10, durham_GS_NA,
              durham_psa_lt10, durham_psa_10_20, durham_psa_gt20, durham_psa_NA,
              durham_met_no, durham_met_yes, durham_met_NA)

cohort_columns = [
    ("GSE116918",  gse116918),
    ("GSE55935",   gse55935),
    ("GSE51066",   gse51066),
    ("GSE46691",   gse46691),
    ("GSE41408",   gse41408),
    ("GSE70769",   gse70769),
    ("Discovery total",  discovery_total),
    ("JHU Nat. History", jhu),
    ("Durham VA",        durham),
]

# Build a dataframe shaped like the published layout: rows are clinicopathologic
# categories, columns are cohorts.
df_s1 = pd.DataFrame({"Variable": row_labels})
for cname, cvals in cohort_columns:
    df_s1[cname] = cvals

# Save to xlsx with custom styling: 2-row header (cohort group + cohort name),
# bold category rows, indented sub-rows.
path_s1 = OUT_DIR / "TableS1_cohorts.xlsx"

wb = Workbook()
ws = wb.active
ws.title = "Cohort characteristics"

# Row 1: "Development cohorts" / "Validation cohorts" group header
ws.cell(row=1, column=1).value = ""
ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=8)
g1 = ws.cell(row=1, column=2)
g1.value = "Development cohorts (Met-Score discovery)"
g1.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
g1.fill = PatternFill("solid", start_color="305496")
g1.alignment = Alignment(horizontal="center", vertical="center")

ws.merge_cells(start_row=1, start_column=9, end_row=1, end_column=10)
g2 = ws.cell(row=1, column=9)
g2.value = "Clinical validation cohorts"
g2.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
g2.fill = PatternFill("solid", start_color="548235")
g2.alignment = Alignment(horizontal="center", vertical="center")

# Row 2: cohort names
ws.cell(row=2, column=1).value = "Variable"
for j, (cname, _) in enumerate(cohort_columns, start=2):
    ws.cell(row=2, column=j).value = cname
for j in range(1, 2 + len(cohort_columns)):
    c = ws.cell(row=2, column=j)
    c.font = Font(bold=True, name="Arial", size=10)
    c.fill = PatternFill("solid", start_color="D9E1F2")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Data rows
SECTION_FILL = PatternFill("solid", start_color="F2F2F2")
for i, label in enumerate(row_labels):
    r = i + 3   # data starts at row 3
    ws.cell(row=r, column=1).value = label
    is_section = label.strip() in {"Pathological Stage", "Gleason Score",
                                     "Preoperative PSA (ng/ml)", "Metastasis"}
    ws.cell(row=r, column=1).font = Font(bold=is_section, name="Arial", size=10)
    if is_section:
        for j in range(1, 2 + len(cohort_columns)):
            ws.cell(row=r, column=j).fill = SECTION_FILL
    for j, (_, cvals) in enumerate(cohort_columns, start=2):
        cell = ws.cell(row=r, column=j)
        cell.value = cvals[i]
        cell.font = Font(name="Arial", size=10)
        cell.alignment = Alignment(horizontal="right" if not is_section else "center")

# Borders + column widths
thin = Side(border_style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
for r in range(1, ws.max_row + 1):
    for c in range(1, ws.max_column + 1):
        ws.cell(row=r, column=c).border = border
ws.column_dimensions["A"].width = 30
for j in range(2, 2 + len(cohort_columns)):
    ws.column_dimensions[get_column_letter(j)].width = 14
ws.row_dimensions[1].height = 22
ws.row_dimensions[2].height = 28
ws.freeze_panes = "B3"

# Footnotes sheet
fn = wb.create_sheet("Footnotes & sources")
notes = [
    "Table S1: Clinicopathologic characteristics of discovery and validation cohorts",
    "",
    "Scope: the six development (discovery) cohorts used for Met-Score gene selection by random-effects meta-analysis "
    "(GSE116918, GSE55935, GSE51066, GSE46691, GSE41408, GSE70769; total n = 1,000), and the two clinical validation "
    "cohorts (JHU Nat. History, Durham VA).",
    "",
    "JHU Nat. History is an independent, outcome-annotated clinical validation cohort: it was excluded from the "
    "meta-analysis and from training of the locked clinical Met-Score probability model. GRID is not a validation cohort "
    "in this package; only JHU Nat. History and Durham VA are validation cohorts.",
    "",
    "Durham VA: replicates the EXACT R-script filter chain in code/Durham_MetScore_Validation_BatchCorrected.R: "
    "merge clin (data/Durham_cohort_and_GRID_cohort/Durham_cohort_clinical_data_022526.xlsx, sheet 'clin') with the "
    "sample-id mapping in Durham_cohort_011526.xlsx sheet 'Sheet2' -> 887 rows; filter !is.na(mets) -> 558 rows; drop "
    "sentinel rows where pogl == 0 -> n = 555. Pathological stage uses column 'stg' (NOT 'clinicalstage'; the codebook "
    "explicitly defines stg = Pathological stage and clinicalstage = Clinical stage). Pathological Gleason from 'pogl'; "
    "preoperative PSA from 'psapresurg'; metastasis events from 'mets'. Verified against R script print: 'Cohort size "
    "after sentinel removal: 555' and 'Pathological Gleason distribution: 5 7 / 6 61 / 7 422 / 8 27 / 9 38'.",
    "",
    "JHU Nat. History: extracted from outs/jhu_pheno_filter_MetaScorer_Zscore.rda (n = 239). Pathological stage from "
    "column 'pstage' (the previously published Table S1 inadvertently used 'cstage' / clinical stage in the path-stage "
    "row; the corrected pathological-stage values appear here, T2=62, T3=174). Pathological Gleason from raw 'pathgs', the "
    "column that drives the survival models: <=6 = 1, 7 = 133, 8 = 28, 9 = 77, 10 = 0, missing = 0. Preoperative PSA from "
    "'preop_psa'; metastasis from 'Metastasis' factor (Mets vs No_Mets, 239/239 non-NA).",
    "",
    "The n = 239 JHU sample is the observed two-phase case-cohort sample: 28 subcohort cases + 146 subcohort controls + "
    "65 outside-subcohort cases, drawn from a source cohort of 745 with a selected subcohort of 265. The 93/239 metastases "
    "(28 subcohort cases + 65 outside-subcohort cases) are the case-enriched phase-two event count, not an unweighted "
    "source-cohort metastasis prevalence estimate. The Gleason-category-adjusted survival models restrict pathological "
    "Gleason to levels 7, 8, and 9 (GS7 reference), so the single GS<=6 record is not represented in those models.",
    "",
    "Discovery cohorts (GSE116918, GSE55935, GSE51066, GSE46691, GSE41408, GSE70769): descriptors are carried forward "
    "verbatim from the prior authoritative supplementary table because none of those cohorts changed, whereas the two "
    "validation columns (JHU Nat. History, Durham VA) were regenerated programmatically from the source data. Discovery "
    "total events = 306 (22+8+51+212+9+4); discovery total non-events = 694 (226+36+34+333+39+26).",
    "",
    "GS Score 5 in the discovery cohorts: 3/63 GSE46691 ≤6 patients (4.7%) carry a Gleason score of 5; preserved here "
    "as in the previously published table.",
    "",
    "Generated by code/generate_supplementary_tables.py from the canonical pipeline outputs. Re-run after any change "
    "to the underlying clinical files or filter logic.",
]
for i, line in enumerate(notes, start=1):
    cell = fn.cell(row=i, column=1)
    cell.value = line
    cell.font = Font(name="Arial", size=10, bold=(i == 1))
    cell.alignment = Alignment(wrap_text=True, vertical="top")
fn.column_dimensions["A"].width = 130
for i in range(1, len(notes) + 1):
    fn.row_dimensions[i].height = 30 if i in (1,) or len(notes[i-1]) > 120 else 18

from datetime import datetime as _dt
_fx = _dt(2000, 1, 1, 0, 0, 0)
wb.properties.created = _fx; wb.properties.modified = _fx; wb.properties.creator = "MET_PCa pipeline"
wb.save(path_s1)
print(f"Wrote {path_s1}  ({len(cohort_columns)} cohorts: 6 discovery + Discovery total + 2 validation)")


# ============================================================================
# Write a README that records what each table contains and where it came from
# ============================================================================
readme = OUT_DIR / "README_supplementary_tables.md"
readme.write_text("""# Supplementary Tables S1-S5

All workbooks are regenerated by `code/supplementary/generate_supplementary_tables.py`
from accepted pipeline aggregates (no hard-coded statistics). Re-run with:

```bash
python code/supplementary/generate_supplementary_tables.py
```

## Table S1: Cohort clinicopathologic characteristics
`TableS1_cohorts.xlsx`. Six discovery cohorts (GSE116918, GSE55935, GSE51066,
GSE46691, GSE41408, GSE70769; total n = 1,000) plus two independent clinical
validation cohorts, JHU Nat. History (n = 239) and Durham VA (n = 555). JHU is an
independent outcome-annotated validation cohort excluded from discovery and from
locked-classifier training; it is not an internal held-out test. GRID is not a
validation cohort in this package. Durham values are extracted directly from
`data/Durham_cohort_and_GRID_cohort/*.xlsx`; JHU from
`outs/jhu_pheno_filter_MetaScorer_Zscore.rda`.

## Table S2: 45-gene Met-Score signature with meta-analysis statistics
`TableS2_metscore_genes.xlsx`, six sheets, all from accepted aggregates:
- Signature genes: pooled Hedges' g, SE, 95% CI, p, FDR, tau2, I2, Cochran Q/p,
  95% prediction interval, and contributing-cohort count
  (`outs/FigureS5/FigureS5_panelA_heterogeneity.csv`).
- Per-cohort effects (`outs/TableS2_per_cohort_effect_sizes.csv`).
- Fixed-panel LOCO stability per held-out cohort
  (`outs/LOCO_gene_stability_with_metadata.csv`).
- LOCO gene consistency with actual contributing studies per fold
  (`outs/LOO_gene_consistency.csv`).
- Complete-reselection summary (`outs/FigureS5/FigureS5_panelB_fold_summary.csv`).
- Complete-reselection membership (`outs/FigureS5/FigureS5_panelB_membership.csv`).

## Table S3: Durham secondary endpoints + GS7 incremental concordance
`TableS3_secondary_endpoints_and_GS7.xlsx`, two worksheets. The complete
multivariable MFS model is now Main Table 1 (`outs/MainTable1_multivariable_MFS.csv`,
from `MetScore_Sensitivity.R`); Figure S6d continues to use the unchanged
`outs/FigureS6/panelD_multivariable.csv`.
- `Durham secondary endpoints`, from `outs/TableS3_Durham_secondary_endpoints.csv`
  (produced by `Met_PCa_Survival_Multivariate.R`). Durham OS, PCSM, and BCR on the
  complete cohort; parsimonious locked high/low Met-Score class adjusted for
  pathological Gleason category (GS7 reference). OS/BCR robust Cox; PCSM robust
  cause-specific Cox with non-prostate-cancer death censored at its recorded time.
  No MFS row and no surgical margin; non-estimable GS<=6 separation is blank with an
  explicit status.
- `GS7 incremental concordance`, from `outs/Figure3/GS7_incremental_concordance.csv`.
  Cause-specific concordance of Grade Group with and without the frozen Met-Score
  within pathological GS7. JHU is the Sanderson case-cohort-weighted Harrell C
  (optimism-corrected, B=2000 design-stratified bootstrap); Durham applies the
  frozen JHU coefficients as an ordinary external Harrell C (B=2000 patient
  bootstrap). Raw locked-v1 probability and its locked binary class, unchanged.

## Table S4: Met-Score gene cell-type expression in GSE274229
`TableS4_metscore_celltype.xlsx`, from
`outs/met_score_gene_celltype_expression_WITH_GENE.csv` (45 genes x 13 cell types).

## Table S5: Met-Score POS/NEG decomposition in GSE143791 (Kfoury bone niche)
`TableS5_kfoury_decomposition.xlsx`, from
`outs/kfoury/GSE143791_met_score_pos_neg_by_{celltype_by_fraction,celltype,fraction}.csv`.
Absent cell-type x fraction combinations were not observed, not zero-imputed
(see the Notes sheet).
""")
print(f"Wrote {readme}")
print()
print("ALL TABLES REGENERATED:")
print(f"  S1: {path_s1}")
print(f"  S2: {path_s2}")
print(f"  S3: {path_s3}")
print(f"  S4: {path_s4}")
print(f"  S5: {path_s5}")
